# # # from flask import Flask, request, jsonify, send_file
# # # from flask_cors import CORS
# # # import pandas as pd
# # # import os
# # # from datetime import datetime
# # # import uuid
# # # from openpyxl import load_workbook
# # # from openpyxl.styles import Font, PatternFill, Alignment
# # # import json

# # # app = Flask(__name__)
# # # CORS(app)

# # # # Configuration
# # # EXCEL_DIR = 'excel_files'
# # # PRIMARY_SHEET = 'primary_applications.xlsx'
# # # os.makedirs(EXCEL_DIR, exist_ok=True)

# # # # Helper function to ensure primary sheet exists
# # # def ensure_primary_sheet():
# # #     primary_path = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
# # #     if not os.path.exists(primary_path):
# # #         df = pd.DataFrame(columns=[
# # #             'Name', 'Address', 'Email', 'Role', 'Experience', 
# # #             'Location', 'Github', 'LinkedIn', 'Personal Sheet Path', 
# # #             'Created At'
# # #         ])
# # #         df.to_excel(primary_path, index=False)
        
# # #         # Style the header
# # #         wb = load_workbook(primary_path)
# # #         ws = wb.active
# # #         for cell in ws[1]:
# # #             cell.font = Font(bold=True)
# # #             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
# # #             cell.font = Font(bold=True, color="FFFFFF")
# # #         wb.save(primary_path)
# # #     return primary_path

# # # # Helper function to style Excel headers
# # # def style_excel_headers(file_path):
# # #     wb = load_workbook(file_path)
# # #     ws = wb.active
# # #     for cell in ws[1]:
# # #         cell.font = Font(bold=True)
# # #         cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
# # #         cell.font = Font(bold=True, color="FFFFFF")
# # #         cell.alignment = Alignment(horizontal='center', vertical='center')
    
# # #     # Auto-adjust column widths
# # #     for column in ws.columns:
# # #         max_length = 0
# # #         column_letter = column[0].column_letter
# # #         for cell in column:
# # #             try:
# # #                 if len(str(cell.value)) > max_length:
# # #                     max_length = len(str(cell.value))
# # #             except:
# # #                 pass
# # #         adjusted_width = min(max_length + 2, 50)
# # #         ws.column_dimensions[column_letter].width = adjusted_width
    
# # #     wb.save(file_path)

# # # @app.route('/health', methods=['GET'])
# # # def health_check():
# # #     """Health check endpoint"""
# # #     return jsonify({'status': 'healthy', 'service': 'Job Application Excel API'}), 200

# # # @app.route('/api/primary-record', methods=['POST'])
# # # def create_primary_record():
# # #     """
# # #     Create or update primary application record
# # #     Expected JSON:
# # #     {
# # #         "Name": "John Doe",
# # #         "Address": "123 Main St",
# # #         "Email": "john@example.com",
# # #         "Role": "Software Developer",
# # #         "Experience": "3 years",
# # #         "Location": "New York",
# # #         "Github": "github.com/john",
# # #         "LinkedIn": "linkedin.com/in/john",
# # #         "Personal Sheet Path": "john_doe_jobs.xlsx"
# # #     }
# # #     """
# # #     try:
# # #         data = request.json
# # #         primary_path = ensure_primary_sheet()
        
# # #         # Read existing data
# # #         df = pd.read_excel(primary_path)
        
# # #         # Check if record exists (by Name or Email)
# # #         name = data.get('Name')
# # #         email = data.get('Email')
        
# # #         existing_index = None
# # #         if not df.empty:
# # #             if name and 'Name' in df.columns:
# # #                 existing_index = df[df['Name'] == name].index
# # #             if existing_index is None or len(existing_index) == 0:
# # #                 if email and 'Email' in df.columns:
# # #                     existing_index = df[df['Email'] == email].index
        
# # #         # Prepare record
# # #         record = {
# # #             'Name': data.get('Name', ''),
# # #             'Address': data.get('Address', ''),
# # #             'Email': data.get('Email', ''),
# # #             'Role': data.get('Role', ''),
# # #             'Experience': data.get('Experience', ''),
# # #             'Location': data.get('Location', ''),
# # #             'Github': data.get('Github', ''),
# # #             'LinkedIn': data.get('LinkedIn', ''),
# # #             'Personal Sheet Path': data.get('Personal Sheet Path', ''),
# # #             'Created At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# # #         }
        
# # #         if existing_index is not None and len(existing_index) > 0:
# # #             # Update existing record
# # #             for key, value in record.items():
# # #                 if key != 'Created At':  # Don't update creation time
# # #                     df.loc[existing_index[0], key] = value
# # #             action = 'updated'
# # #         else:
# # #             # Add new record
# # #             df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
# # #             action = 'created'
        
# # #         # Save
# # #         df.to_excel(primary_path, index=False)
# # #         style_excel_headers(primary_path)
        
# # #         return jsonify({
# # #             'success': True,
# # #             'action': action,
# # #             'record': record,
# # #             'file_path': primary_path
# # #         }), 200
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/create-personal-sheet', methods=['POST'])
# # # def create_personal_sheet():
# # #     """
# # #     Create a new personal job tracking sheet
# # #     Expected JSON:
# # #     {
# # #         "name": "John Doe"
# # #     }
# # #     Returns the sheet ID/path
# # #     """
# # #     try:
# # #         data = request.json
# # #         name = data.get('name', 'Unknown')
        
# # #         # Create filename
# # #         safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
# # #         safe_name = safe_name.replace(' ', '_')
# # #         sheet_id = f"{safe_name}_{uuid.uuid4().hex[:8]}"
# # #         filename = f"{sheet_id}_jobs.xlsx"
# # #         filepath = os.path.join(EXCEL_DIR, filename)
        
# # #         # Create sheet with headers
# # #         df = pd.DataFrame(columns=[
# # #             'Job Title', 'Company Name', 'Job Skills', 
# # #             'Job Description', 'Job Link', 'Updated Resume Link',
# # #             'Applied Date', 'Status'
# # #         ])
# # #         df.to_excel(filepath, index=False)
# # #         style_excel_headers(filepath)
        
# # #         return jsonify({
# # #             'success': True,
# # #             'sheet_id': sheet_id,
# # #             'filename': filename,
# # #             'filepath': filepath
# # #         }), 200
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/append-jobs', methods=['POST'])
# # # def append_jobs():
# # #     """
# # #     Append job listings to personal sheet
# # #     Expected JSON:
# # #     {
# # #         "sheet_id": "john_doe_abc123",
# # #         "jobs": [
# # #             {
# # #                 "Job Title": "Software Developer",
# # #                 "Company Name": "Tech Corp",
# # #                 "Job Skills": "Python, React",
# # #                 "Job Description": "...",
# # #                 "Job Link": "https://..."
# # #             }
# # #         ]
# # #     }
# # #     """
# # #     try:
# # #         data = request.json
# # #         sheet_id = data.get('sheet_id')
# # #         jobs = data.get('jobs', [])
        
# # #         if not sheet_id:
# # #             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400
        
# # #         # Find the file
# # #         filename = f"{sheet_id}_jobs.xlsx"
# # #         filepath = os.path.join(EXCEL_DIR, filename)
        
# # #         if not os.path.exists(filepath):
# # #             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
# # #         # Read existing data
# # #         df = pd.read_excel(filepath)
        
# # #         # Append jobs
# # #         for job in jobs:
# # #             job_record = {
# # #                 'Job Title': job.get('Job Title', ''),
# # #                 'Company Name': job.get('Company Name', ''),
# # #                 'Job Skills': job.get('Job Skills', ''),
# # #                 'Job Description': job.get('Job Description', ''),
# # #                 'Job Link': job.get('Job Link', ''),
# # #                 'Updated Resume Link': job.get('Updated Resume Link', ''),
# # #                 'Applied Date': datetime.now().strftime('%Y-%m-%d'),
# # #                 'Status': job.get('Status', 'Pending')
# # #             }
# # #             df = pd.concat([df, pd.DataFrame([job_record])], ignore_index=True)
        
# # #         # Save
# # #         df.to_excel(filepath, index=False)
# # #         style_excel_headers(filepath)
        
# # #         return jsonify({
# # #             'success': True,
# # #             'jobs_added': len(jobs),
# # #             'filepath': filepath
# # #         }), 200
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/update-job-resume', methods=['POST'])
# # # def update_job_resume():
# # #     """
# # #     Update specific job with resume link
# # #     Expected JSON:
# # #     {
# # #         "sheet_id": "john_doe_abc123",
# # #         "job_link": "https://...",
# # #         "resume_link": "https://..."
# # #     }
# # #     """
# # #     try:
# # #         data = request.json
# # #         sheet_id = data.get('sheet_id')
# # #         job_link = data.get('job_link')
# # #         resume_link = data.get('resume_link')
        
# # #         if not sheet_id or not job_link:
# # #             return jsonify({'success': False, 'error': 'sheet_id and job_link are required'}), 400
        
# # #         # Find the file
# # #         filename = f"{sheet_id}_jobs.xlsx"
# # #         filepath = os.path.join(EXCEL_DIR, filename)
        
# # #         if not os.path.exists(filepath):
# # #             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
# # #         # Read and update
# # #         df = pd.read_excel(filepath)
        
# # #         # Find matching job
# # #         mask = df['Job Link'] == job_link
# # #         if mask.any():
# # #             df.loc[mask, 'Updated Resume Link'] = resume_link
# # #             df.loc[mask, 'Status'] = 'Resume Tailored'
            
# # #             # Save
# # #             df.to_excel(filepath, index=False)
# # #             style_excel_headers(filepath)
            
# # #             return jsonify({
# # #                 'success': True,
# # #                 'updated': True,
# # #                 'filepath': filepath
# # #             }), 200
# # #         else:
# # #             return jsonify({
# # #                 'success': False,
# # #                 'error': 'Job not found'
# # #             }), 404
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/get-jobs', methods=['POST'])
# # # def get_jobs():
# # #     """
# # #     Get all jobs from personal sheet
# # #     Expected JSON:
# # #     {
# # #         "sheet_id": "john_doe_abc123"
# # #     }
# # #     """
# # #     try:
# # #         data = request.json
# # #         sheet_id = data.get('sheet_id')
        
# # #         if not sheet_id:
# # #             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400
        
# # #         # Find the file
# # #         filename = f"{sheet_id}_jobs.xlsx"
# # #         filepath = os.path.join(EXCEL_DIR, filename)
        
# # #         if not os.path.exists(filepath):
# # #             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
# # #         # Read data
# # #         df = pd.read_excel(filepath)
        
# # #         # Convert to list of dictionaries
# # #         jobs = df.to_dict('records')
        
# # #         return jsonify({
# # #             'success': True,
# # #             'jobs': jobs,
# # #             'count': len(jobs)
# # #         }), 200
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/download/<sheet_id>', methods=['GET'])
# # # def download_sheet(sheet_id):
# # #     """Download Excel file"""
# # #     try:
# # #         filename = f"{sheet_id}_jobs.xlsx"
# # #         filepath = os.path.join(EXCEL_DIR, filename)
        
# # #         if not os.path.exists(filepath):
# # #             # Try primary sheet
# # #             if sheet_id == 'primary':
# # #                 filepath = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
# # #             else:
# # #                 return jsonify({'success': False, 'error': 'File not found'}), 404
        
# # #         return send_file(filepath, as_attachment=True, download_name=filename)
        
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # @app.route('/api/list-sheets', methods=['GET'])
# # # def list_sheets():
# # #     """List all Excel sheets"""
# # #     try:
# # #         files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]
# # #         return jsonify({
# # #             'success': True,
# # #             'files': files,
# # #             'count': len(files)
# # #         }), 200
# # #     except Exception as e:
# # #         return jsonify({'success': False, 'error': str(e)}), 500

# # # if __name__ == '__main__':
# # #     app.run(host='0.0.0.0', port=5001, debug=True)
# # from flask import Flask, request, jsonify, send_file
# # from flask_cors import CORS
# # import pandas as pd
# # import os
# # from datetime import datetime
# # import uuid
# # from openpyxl import load_workbook
# # from openpyxl.styles import Font, PatternFill, Alignment
# # import json

# # app = Flask(__name__)
# # CORS(app)

# # # Configuration
# # EXCEL_DIR = 'excel_files'
# # PRIMARY_SHEET = 'primary_applications.xlsx'
# # os.makedirs(EXCEL_DIR, exist_ok=True)

# # # Helper function to parse incoming data
# # def parse_request_data():
# #     """
# #     Parse incoming request data, handling both JSON objects and JSON strings
# #     """
# #     try:
# #         data = request.get_json()
        
# #         # If data is a string, parse it
# #         if isinstance(data, str):
# #             data = json.loads(data)
        
# #         return data
        
# #     except Exception as e:
# #         return None

# # # Helper function to ensure primary sheet exists
# # def ensure_primary_sheet():
# #     primary_path = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
# #     if not os.path.exists(primary_path):
# #         df = pd.DataFrame(columns=[
# #             'Name', 'Address', 'Email', 'Role', 'Experience', 
# #             'Location', 'Github', 'LinkedIn', 'Personal Sheet Path', 
# #             'Created At'
# #         ])
# #         df.to_excel(primary_path, index=False)
        
# #         # Style the header
# #         wb = load_workbook(primary_path)
# #         ws = wb.active
# #         for cell in ws[1]:
# #             cell.font = Font(bold=True)
# #             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
# #             cell.font = Font(bold=True, color="FFFFFF")
# #         wb.save(primary_path)
# #     return primary_path

# # # Helper function to style Excel headers
# # def style_excel_headers(file_path):
# #     wb = load_workbook(file_path)
# #     ws = wb.active
# #     for cell in ws[1]:
# #         cell.font = Font(bold=True)
# #         cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
# #         cell.font = Font(bold=True, color="FFFFFF")
# #         cell.alignment = Alignment(horizontal='center', vertical='center')
    
# #     # Auto-adjust column widths
# #     for column in ws.columns:
# #         max_length = 0
# #         column_letter = column[0].column_letter
# #         for cell in column:
# #             try:
# #                 if len(str(cell.value)) > max_length:
# #                     max_length = len(str(cell.value))
# #             except:
# #                 pass
# #         adjusted_width = min(max_length + 2, 50)
# #         ws.column_dimensions[column_letter].width = adjusted_width
    
# #     wb.save(file_path)

# # @app.route('/health', methods=['GET'])
# # def health_check():
# #     """Health check endpoint"""
# #     return jsonify({'status': 'healthy', 'service': 'Job Application Excel API'}), 200

# # @app.route('/api/primary-record', methods=['POST'])
# # def create_primary_record():
# #     """
# #     Create or update primary application record
# #     Expected JSON:
# #     {
# #         "Name": "John Doe",
# #         "Address": "123 Main St",
# #         "Email": "john@example.com",
# #         "Role": "Software Developer",
# #         "Experience": "3 years",
# #         "Location": "New York",
# #         "Github": "github.com/john",
# #         "LinkedIn": "linkedin.com/in/john",
# #         "Personal Sheet Path": "john_doe_jobs.xlsx"
# #     }
# #     """
# #     try:
# #         data = parse_request_data()
        
# #         primary_path = ensure_primary_sheet()
        
# #         # Read existing data
# #         df = pd.read_excel(primary_path)
        
# #         # Check if record exists (by Name or Email)
# #         name = data.get('Name')
# #         email = data.get('Email')
        
# #         existing_index = None
# #         if not df.empty:
# #             if name and 'Name' in df.columns:
# #                 existing_index = df[df['Name'] == name].index
# #             if existing_index is None or len(existing_index) == 0:
# #                 if email and 'Email' in df.columns:
# #                     existing_index = df[df['Email'] == email].index
        
# #         # Prepare record
# #         record = {
# #             'Name': data.get('Name', ''),
# #             'Address': data.get('Address', ''),
# #             'Email': data.get('Email', ''),
# #             'Role': data.get('Role', ''),
# #             'Experience': data.get('Experience', ''),
# #             'Location': data.get('Location', ''),
# #             'Github': data.get('Github', ''),
# #             'LinkedIn': data.get('LinkedIn', ''),
# #             'Personal Sheet Path': data.get('Personal Sheet Path', ''),
# #             'Created At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
# #         }
        
# #         if existing_index is not None and len(existing_index) > 0:
# #             # Update existing record
# #             for key, value in record.items():
# #                 if key != 'Created At':  # Don't update creation time
# #                     df.loc[existing_index[0], key] = value
# #             action = 'updated'
# #         else:
# #             # Add new record
# #             df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
# #             action = 'created'
        
# #         # Save
# #         df.to_excel(primary_path, index=False)
# #         style_excel_headers(primary_path)
        
# #         return jsonify({
# #             'success': True,
# #             'action': action,
# #             'record': record,
# #             'file_path': primary_path
# #         }), 200
        
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # @app.route('/api/create-personal-sheet', methods=['POST'])
# # def create_personal_sheet():
# #     """
# #     Create a new personal job tracking sheet
# #     Expected JSON:
# #     {
# #         "name": "John Doe"
# #     }
# #     Returns the sheet ID/path
# #     """
# #     try:
# #         data = parse_request_data()
# #         name = data.get('name', 'Unknown')
        
# #         # Create filename
# #         safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
# #         safe_name = safe_name.replace(' ', '_')
# #         sheet_id = f"{safe_name}_{uuid.uuid4().hex[:8]}"
# #         filename = f"{sheet_id}_jobs.xlsx"
# #         filepath = os.path.join(EXCEL_DIR, filename)
        
# #         # Create sheet with headers
# #         df = pd.DataFrame(columns=[
# #             'Job Title', 'Company Name', 'Job Skills', 
# #             'Job Description', 'Job Link', 'Updated Resume Link',
# #             'Applied Date', 'Status'
# #         ])
# #         df.to_excel(filepath, index=False)
# #         style_excel_headers(filepath)
        
# #         return jsonify({
# #             'success': True,
# #             'sheet_id': sheet_id,
# #             'filename': filename,
# #             'filepath': filepath
# #         }), 200
        
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # from openpyxl import Workbook, load_workbook
# # import os

# # @app.route("/api/append-jobs", methods=["POST"])
# # def append_jobs():
# #     """
# #     Append job listings to personal sheet
# #     Expected JSON:
# #     {
# #         "sheet_id": "john_doe_abc123",
# #         "jobs": [
# #             {
# #                 "Job Title": "Software Developer",
# #                 "Company Name": "Tech Corp",
# #                 "Job Skills": "Python, React",
# #                 "Job Description": "...",
# #                 "Job Link": "https://..."
# #             }
# #         ]
# #     }
# #     """
# #     try:
# #         data = parse_request_data()
        
# #         if not data:
# #             return jsonify({'success': False, 'error': 'Invalid JSON data'}), 400
        
# #         # Get sheet_id - try multiple sources
# #         sheet_id = data.get('sheet_id')
        
# #         # If sheet_id is not at root level, try to get it from the first job
# #         if not sheet_id and 'jobs' in data and len(data['jobs']) > 0:
# #             sheet_id = data['jobs'][0].get('sheet_id')
        
# #         jobs = data.get('jobs', [])
        
# #         if not sheet_id:
# #             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400
        
# #         if not jobs or len(jobs) == 0:
# #             return jsonify({'success': False, 'error': 'No jobs provided'}), 400
        
# #         # Correct file path with EXCEL_DIR and proper naming
# #         filename = f"{sheet_id}_jobs.xlsx"
# #         filepath = os.path.join(EXCEL_DIR, filename)
        
# #         # Check if file exists
# #         if not os.path.exists(filepath):
# #             # Create new file with proper structure
# #             df = pd.DataFrame(columns=[
# #                 'Job Title', 'Company Name', 'Job Skills', 
# #                 'Job Description', 'Job Link', 'Updated Resume Link',
# #                 'Applied Date', 'Status'
# #             ])
# #             df.to_excel(filepath, index=False)
# #             style_excel_headers(filepath)
        
# #         # Read existing data
# #         try:
# #             df = pd.read_excel(filepath)
# #         except Exception as e:
# #             return jsonify({
# #                 'success': False, 
# #                 'error': f'Failed to read Excel file: {str(e)}'
# #             }), 500
        
# #         # Append jobs
# #         jobs_added = 0
# #         for job in jobs:
# #             # Clean the job data - remove sheet_id if it exists in individual jobs
# #             job_record = {
# #                 'Job Title': job.get('Job Title', ''),
# #                 'Company Name': job.get('Company Name', ''),
# #                 'Job Skills': job.get('Job Skills', ''),
# #                 'Job Description': job.get('Job Description', ''),
# #                 'Job Link': job.get('Job Link', ''),
# #                 'Updated Resume Link': '',
# #                 'Applied Date': datetime.now().strftime('%Y-%m-%d'),
# #                 'Status': 'Pending'
# #             }
# #             df = pd.concat([df, pd.DataFrame([job_record])], ignore_index=True)
# #             jobs_added += 1
        
# #         # Save
# #         df.to_excel(filepath, index=False)
# #         style_excel_headers(filepath)
        
# #         return jsonify({
# #             'success': True,
# #             'message': 'Jobs appended successfully',
# #             'jobs_added': jobs_added,
# #             'filepath': filepath,
# #             'sheet_id': sheet_id
# #         }), 200
        
# #     except Exception as e:
# #         import traceback
# #         return jsonify({
# #             'success': False, 
# #             'error': f'Error in append_jobs: {str(e)}',
# #             'traceback': traceback.format_exc()
# #         }), 500

# # @app.route('/api/update-job-resume', methods=['POST'])
# # def update_job_resume():
# #     """
# #     Update specific job with resume link
# #     Expected JSON:
# #     {
# #         "sheet_id": "john_doe_abc123",
# #         "job_link": "https://...",
# #         "resume_link": "https://..."
# #     }
# #     """
# #     try:
# #         data = parse_request_data()
        
# #         sheet_id = data.get('sheet_id')
# #         job_link = data.get('job_link')
# #         resume_link = data.get('resume_link')
        
# #         if not sheet_id or not job_link:
# #             return jsonify({'success': False, 'error': 'sheet_id and job_link are required'}), 400
        
# #         # Find the file
# #         filename = f"{sheet_id}_jobs.xlsx"
# #         filepath = os.path.join(EXCEL_DIR, filename)
        
# #         if not os.path.exists(filepath):
# #             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
# #         # Read and update
# #         df = pd.read_excel(filepath)
        
# #         # Find matching job
# #         mask = df['Job Link'] == job_link
# #         if mask.any():
# #             df.loc[mask, 'Updated Resume Link'] = resume_link
# #             df.loc[mask, 'Status'] = 'Resume Tailored'
            
# #             # Save
# #             df.to_excel(filepath, index=False)
# #             style_excel_headers(filepath)
            
# #             return jsonify({
# #                 'success': True,
# #                 'updated': True,
# #                 'filepath': filepath
# #             }), 200
# #         else:
# #             return jsonify({
# #                 'success': False,
# #                 'error': 'Job not found'
# #             }), 404
        
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # @app.route('/api/get-jobs', methods=['POST'])
# # def get_jobs():
# #     """
# #     Get all jobs from personal sheet
# #     Expected JSON:
# #     {
# #         "sheet_id": "john_doe_abc123"
# #     }
# #     """
# #     try:
# #         data = parse_request_data()
        
# #         sheet_id = data.get('sheet_id')
        
# #         if not sheet_id:
# #             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400
        
# #         # Find the file
# #         filename = f"{sheet_id}_jobs.xlsx"
# #         filepath = os.path.join(EXCEL_DIR, filename)
        
# #         if not os.path.exists(filepath):
# #             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
# #         # Read data
# #         df = pd.read_excel(filepath)
        
# #         # Convert to list of dictionaries
# #         jobs = df.to_dict('records')
        
# #         return jsonify({
# #             'success': True,
# #             'jobs': jobs,
# #             'count': len(jobs)
# #         }), 200
        
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # @app.route('/api/download/<sheet_id>', methods=['GET'])
# # def download_sheet(sheet_id):
# #     """Download Excel file"""
# #     try:
# #         filename = f"{sheet_id}_jobs.xlsx"
# #         filepath = os.path.join(EXCEL_DIR, filename)
        
# #         if not os.path.exists(filepath):
# #             # Try primary sheet
# #             if sheet_id == 'primary':
# #                 filepath = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
# #             else:
# #                 return jsonify({'success': False, 'error': 'File not found'}), 404
        
# #         return send_file(filepath, as_attachment=True, download_name=filename)
        
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # @app.route('/api/list-sheets', methods=['GET'])
# # def list_sheets():
# #     """List all Excel sheets"""
# #     try:
# #         files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]
# #         return jsonify({
# #             'success': True,
# #             'files': files,
# #             'count': len(files)
# #         }), 200
# #     except Exception as e:
# #         return jsonify({'success': False, 'error': str(e)}), 500

# # if __name__ == '__main__':
# #     app.run(host='0.0.0.0', port=5001, debug=True)

# from flask import Flask, request, jsonify, send_file
# from flask_cors import CORS
# import pandas as pd
# import os
# from datetime import datetime
# import uuid
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment
# import json
# import time

# app = Flask(__name__)
# CORS(app)

# # Configuration
# EXCEL_DIR = os.path.join(os.path.dirname(__file__), 'excel_files')
# PRIMARY_SHEET = 'primary_applications.xlsx'


# # Helper function to parse incoming data
# def parse_request_data():
#     """
#     Parse incoming request data, handling both JSON objects and JSON strings
#     """
#     try:
#         data = request.get_json()
        
#         # If data is a string, parse it
#         if isinstance(data, str):
#             data = json.loads(data)
        
#         return data
        
#     except Exception as e:
#         return None

# # Helper function to ensure primary sheet exists
# def ensure_primary_sheet():
#     primary_path = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
#     if not os.path.exists(primary_path):
#         # Create with openpyxl directly
#         wb = Workbook()
#         ws = wb.active
        
#         headers = ['Name', 'Address', 'Email', 'Role', 'Experience', 
#                    'Location', 'Github', 'LinkedIn', 'Personal Sheet Path', 'Created At']
#         ws.append(headers)
        
#         # Style headers
#         for cell in ws[1]:
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center', vertical='center')
        
#         wb.save(primary_path)
#         wb.close()
#     return primary_path

# # Helper function to style Excel headers using openpyxl
# def style_excel_headers(file_path, retries=3, delay=0.5):
#     """
#     Style Excel headers with retry logic to handle file lock issues
#     """
#     for attempt in range(retries):
#         try:
#             # Wait a bit before trying to open
#             time.sleep(delay)
            
#             wb = load_workbook(file_path)
#             ws = wb.active
            
#             # Style headers
#             for cell in ws[1]:
#                 cell.font = Font(bold=True, color="FFFFFF")
#                 cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#                 cell.alignment = Alignment(horizontal='center', vertical='center')
            
#             # Auto-adjust column widths
#             for column in ws.columns:
#                 max_length = 0
#                 column_letter = column[0].column_letter
#                 for cell in column:
#                     try:
#                         if cell.value and len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = min(max_length + 2, 50)
#                 ws.column_dimensions[column_letter].width = adjusted_width
            
#             wb.save(file_path)
#             wb.close()
#             return True
            
#         except Exception as e:
#             if attempt < retries - 1:
#                 time.sleep(delay * (attempt + 1))
#                 continue
#             else:
#                 print(f"Failed to style headers after {retries} attempts: {str(e)}")
#                 return False

# @app.route('/health', methods=['GET'])
# def health_check():
#     """Health check endpoint"""
#     return jsonify({'status': 'healthy', 'service': 'Job Application Excel API'}), 200

# @app.route('/api/primary-record', methods=['POST'])
# def create_primary_record():
#     """
#     Create or update primary application record
#     """
#     try:
#         data = parse_request_data()
        
#         primary_path = ensure_primary_sheet()
        
#         # Read existing data
#         df = pd.read_excel(primary_path)
        
#         # Check if record exists (by Name or Email)
#         name = data.get('Name')
#         email = data.get('Email')
        
#         existing_index = None
#         if not df.empty:
#             if name and 'Name' in df.columns:
#                 existing_index = df[df['Name'] == name].index
#             if existing_index is None or len(existing_index) == 0:
#                 if email and 'Email' in df.columns:
#                     existing_index = df[df['Email'] == email].index
        
#         # Prepare record
#         record = {
#             'Name': data.get('Name', ''),
#             'Address': data.get('Address', ''),
#             'Email': data.get('Email', ''),
#             'Role': data.get('Role', ''),
#             'Experience': data.get('Experience', ''),
#             'Location': data.get('Location', ''),
#             'Github': data.get('Github', ''),
#             'LinkedIn': data.get('LinkedIn', ''),
#             'Personal Sheet Path': data.get('Personal Sheet Path', ''),
#             'Created At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#         }
        
#         if existing_index is not None and len(existing_index) > 0:
#             # Update existing record
#             for key, value in record.items():
#                 if key != 'Created At':  # Don't update creation time
#                     df.loc[existing_index[0], key] = value
#             action = 'updated'
#         else:
#             # Add new record
#             df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
#             action = 'created'
        
#         # Save
#         df.to_excel(primary_path, index=False)
#         style_excel_headers(primary_path)
        
#         return jsonify({
#             'success': True,
#             'action': action,
#             'record': record,
#             'file_path': primary_path
#         }), 200
        
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route('/api/create-personal-sheet', methods=['POST'])
# def create_personal_sheet():
#     """
#     Create a new personal job tracking sheet using openpyxl directly
#     """
#     try:
#         data = parse_request_data()
#         name = data.get('name', 'Unknown')
        
#         # Create filename
#         safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip()
#         safe_name = safe_name.replace(' ', '_')
#         sheet_id = f"{safe_name}_{uuid.uuid4().hex[:8]}"
#         filename = f"{sheet_id}_jobs.xlsx"
#         filepath = os.path.join(EXCEL_DIR, filename)
        
#         # Create with openpyxl to avoid corruption
#         wb = Workbook()
#         ws = wb.active
        
#         headers = ['Job Title', 'Company Name', 'Job Skills', 
#                    'Job Description', 'Job Link', 'Updated Resume Link',
#                    'Applied Date', 'Status']
#         ws.append(headers)
        
#         # Style headers immediately
#         for cell in ws[1]:
#             cell.font = Font(bold=True, color="FFFFFF")
#             cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#             cell.alignment = Alignment(horizontal='center', vertical='center')
        
#         # Set column widths
#         ws.column_dimensions['A'].width = 40  # Job Title
#         ws.column_dimensions['B'].width = 30  # Company Name
#         ws.column_dimensions['C'].width = 50  # Job Skills
#         ws.column_dimensions['D'].width = 50  # Job Description
#         ws.column_dimensions['E'].width = 50  # Job Link
#         ws.column_dimensions['F'].width = 50  # Updated Resume Link
#         ws.column_dimensions['G'].width = 15  # Applied Date
#         ws.column_dimensions['H'].width = 20  # Status
        
#         wb.save(filepath)
#         wb.close()
        
#         return jsonify({
#             'success': True,
#             'sheet_id': sheet_id,
#             'filename': filename,
#             'filepath': filepath
#         }), 200
        
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route("/api/append-jobs", methods=["POST"])
# def append_jobs():
#     """
#     Append job listings to personal sheet using openpyxl to avoid corruption.
#     Includes validation for file integrity and better error handling.
#     """
#     try:
#         data = parse_request_data()

#         if not data:
#             return jsonify({'success': False, 'error': 'Invalid JSON data'}), 400

#         # Get sheet_id
#         sheet_id = data.get('sheet_id')
#         if not sheet_id and 'jobs' in data and len(data['jobs']) > 0:
#             sheet_id = data['jobs'][0].get('sheet_id')

#         jobs = data.get('jobs', [])

#         if not sheet_id:
#             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400

#         if not jobs or len(jobs) == 0:
#             return jsonify({'success': False, 'error': 'No jobs provided'}), 400

#         # File path
#         filename = f"{sheet_id}_jobs.xlsx"
#         filepath = os.path.join(EXCEL_DIR, filename)

#         # Check if file exists
#         if not os.path.exists(filepath):
#             return jsonify({
#                 'success': False,
#                 'error': f'Sheet not found: {filepath}. Please create it first using /api/create-personal-sheet'
#             }), 404

#         # Validate that the file is a real Excel file
#         try:
#             with open(filepath, 'rb') as f:
#                 header = f.read(4)
#                 if header != b'PK\x03\x04':
#                     return jsonify({
#                         'success': False,
#                         'error': f'Corrupted or invalid Excel file at {filepath}. '
#                                  f'Please recreate it using /api/create-personal-sheet.'
#                     }), 400
#         except Exception as e:
#             return jsonify({'success': False, 'error': f'File read error: {str(e)}'}), 500

#         # Safely append jobs
#         try:
#             wb = load_workbook(filepath)
#             ws = wb.active

#             jobs_added = 0
#             for job in jobs:
#                 row_data = [
#                     job.get('Job Title', ''),
#                     job.get('Company Name', ''),
#                     job.get('Job Skills', ''),
#                     job.get('Job Description', ''),
#                     job.get('Job Link', ''),
#                     '',  # Updated Resume Link
#                     datetime.now().strftime('%Y-%m-%d'),  # Applied Date
#                     'Pending'  # Status
#                 ]
#                 ws.append(row_data)
#                 jobs_added += 1

#             wb.save(filepath)
#             wb.close()

#             return jsonify({
#                 'success': True,
#                 'message': 'Jobs appended successfully',
#                 'jobs_added': jobs_added,
#                 'filepath': filepath,
#                 'sheet_id': sheet_id
#             }), 200

#         except Exception as e:
#             return jsonify({
#                 'success': False,
#                 'error': f'Failed to append jobs: {str(e)}'
#             }), 500

#     except Exception as e:
#         import traceback
#         return jsonify({
#             'success': False,
#             'error': f'Error in append_jobs: {str(e)}',
#             'traceback': traceback.format_exc()
#         }), 500

# @app.route('/api/update-job-resume', methods=['POST'])
# def update_job_resume():
#     """
#     Update specific job with resume link
#     """
#     try:
#         data = parse_request_data()
        
#         sheet_id = data.get('sheet_id')
#         job_link = data.get('job_link')
#         resume_link = data.get('resume_link')
        
#         if not sheet_id or not job_link:
#             return jsonify({'success': False, 'error': 'sheet_id and job_link are required'}), 400
        
#         # Find the file
#         filename = f"{sheet_id}_jobs.xlsx"
#         filepath = os.path.join(EXCEL_DIR, filename)
        
#         if not os.path.exists(filepath):
#             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
#         # Read and update
#         df = pd.read_excel(filepath)
        
#         # Find matching job
#         mask = df['Job Link'] == job_link
#         if mask.any():
#             df.loc[mask, 'Updated Resume Link'] = resume_link
#             df.loc[mask, 'Status'] = 'Resume Tailored'
            
#             # Save
#             df.to_excel(filepath, index=False)
#             style_excel_headers(filepath)
            
#             return jsonify({
#                 'success': True,
#                 'updated': True,
#                 'filepath': filepath
#             }), 200
#         else:
#             return jsonify({
#                 'success': False,
#                 'error': 'Job not found'
#             }), 404
        
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route('/api/get-jobs', methods=['POST'])
# def get_jobs():
#     """
#     Get all jobs from personal sheet
#     """
#     try:
#         data = parse_request_data()
        
#         sheet_id = data.get('sheet_id')
        
#         if not sheet_id:
#             return jsonify({'success': False, 'error': 'sheet_id is required'}), 400
        
#         # Find the file
#         filename = f"{sheet_id}_jobs.xlsx"
#         filepath = os.path.join(EXCEL_DIR, filename)
        
#         if not os.path.exists(filepath):
#             return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        
#         # Read data
#         df = pd.read_excel(filepath)
        
#         # Convert to list of dictionaries
#         jobs = df.to_dict('records')
        
#         return jsonify({
#             'success': True,
#             'jobs': jobs,
#             'count': len(jobs)
#         }), 200
        
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route('/api/download/<sheet_id>', methods=['GET'])
# def download_sheet(sheet_id):
#     """Download Excel file"""
#     try:
#         filename = f"{sheet_id}_jobs.xlsx"
#         filepath = os.path.join(EXCEL_DIR, filename)
        
#         if not os.path.exists(filepath):
#             if sheet_id == 'primary':
#                 filepath = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
#             else:
#                 return jsonify({'success': False, 'error': 'File not found'}), 404
        
#         return send_file(filepath, as_attachment=True, download_name=filename)
        
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# @app.route('/api/list-sheets', methods=['GET'])
# def list_sheets():
#     """List all Excel sheets"""
#     try:
#         files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]
#         return jsonify({
#             'success': True,
#             'files': files,
#             'count': len(files)
#         }), 200
#     except Exception as e:
#         return jsonify({'success': False, 'error': str(e)}), 500

# if __name__ == '__main__':
#     app.run(host='0.0.0.0', port=5002, debug=True)
#!/usr/bin/env python3
"""
Flask backend for n8n job-appender flow.
Endpoints:
 - GET  /health
 - POST /api/create-personal-sheet     -> { "name": "Alice" }
 - POST /api/append-jobs              -> accepts n8n-style array of job items OR {"sheet_id":.., "jobs":[...]}
 - POST /api/get-jobs                 -> { "sheet_id": "Alice_abcdef01" }
 - POST /api/update-job-resume        -> { "sheet_id": "..", "job_link": "..", "resume_link": ".." }
 - POST /api/primary-record           -> create/update primary_applications.xlsx record
 - GET  /api/download/<sheet_id>
 - GET  /api/list-sheets
"""
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
from datetime import datetime
import uuid
import json
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

app = Flask(__name__)
CORS(app)

# Configuration
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, 'excel_files')
PRIMARY_SHEET = 'primary_applications.xlsx'

# Ensure excel dir exists
os.makedirs(EXCEL_DIR, exist_ok=True)


# -------------------------
# Utility / helper methods
# -------------------------
def parse_request_data():
    """
    Parse incoming request JSON in a robust way.
    Accepts:
      - JSON object: {"sheet_id": "...", "jobs": [...]}
      - List of job objects (n8n default): [ { "Job Title": "...", "sheet_id": "..." }, {...}, ... ]
      - JSON string body (rare)
    Returns a dict with keys: { 'sheet_id': str|None, 'jobs': list }
    """
    try:
        data = request.get_json(force=True, silent=True)
        if data is None:
            # Try as raw text
            raw = request.data.decode('utf-8') if request.data else ''
            if raw:
                data = json.loads(raw)
            else:
                return {'sheet_id': None, 'jobs': []}
    except Exception:
        # fallback
        try:
            data = json.loads(request.data.decode('utf-8') or '{}')
        except Exception:
            data = None

    # If data is a list (n8n often sends an array of items)
    if isinstance(data, list):
        jobs = []
        sheet_id = None
        for item in data:
            if not isinstance(item, dict):
                continue
            # item may include nested keys like item['json'] in some setups
            candidate = item.get('json') if 'json' in item and isinstance(item['json'], dict) else item
            # collect sheet_id if present
            if not sheet_id:
                sheet_id = candidate.get('sheet_id') or candidate.get('sheetId') or candidate.get('sheetID')
            # convert to normalized job dict expected by this app
            job = {
                'Job Title': candidate.get('Job Title') or candidate.get('job_title') or '',
                'Company Name': candidate.get('Company Name') or candidate.get('company') or '',
                'Job Skills': candidate.get('Job Skills') or candidate.get('job_skills') or '',
                'Job Description': candidate.get('Job Description') or candidate.get('description') or '',
                'Job Link': candidate.get('Job Link') or candidate.get('job_link') or '',
                # preserve original sheet_id if present on each item
                'sheet_id': candidate.get('sheet_id') or candidate.get('sheetId') or None
            }
            jobs.append(job)
        # prefer sheet_id found as top-level or item-level
        return {'sheet_id': sheet_id, 'jobs': jobs}

    # If data is dict
    if isinstance(data, dict):
        # Normalize
        sheet_id = data.get('sheet_id') or data.get('sheetId') or data.get('sheetID')
        jobs = []
        if 'jobs' in data and isinstance(data['jobs'], list):
            # Might be list of dicts already shaped correctly
            for j in data['jobs']:
                if isinstance(j, dict):
                    jobs.append({
                        'Job Title': j.get('Job Title') or j.get('job_title') or '',
                        'Company Name': j.get('Company Name') or j.get('company') or '',
                        'Job Skills': j.get('Job Skills') or j.get('job_skills') or '',
                        'Job Description': j.get('Job Description') or j.get('description') or '',
                        'Job Link': j.get('Job Link') or j.get('job_link') or '',
                        'sheet_id': j.get('sheet_id') or None
                    })
        else:
            # Maybe single job object (n8n configured differently)
            # Try to read keys that look like job fields
            job_keys = {'Job Title', 'Company Name', 'Job Skills', 'Job Description', 'Job Link'}
            if any(k in data for k in job_keys):
                jobs.append({
                    'Job Title': data.get('Job Title', ''),
                    'Company Name': data.get('Company Name', ''),
                    'Job Skills': data.get('Job Skills', ''),
                    'Job Description': data.get('Job Description', ''),
                    'Job Link': data.get('Job Link', ''),
                    'sheet_id': sheet_id
                })

        # If still empty and top-level contains 'items' or 'data'
        if not jobs:
            for alt in ('items', 'data'):
                if alt in data and isinstance(data[alt], list):
                    for item in data[alt]:
                        if isinstance(item, dict):
                            candidate = item.get('json') if 'json' in item else item
                            jobs.append({
                                'Job Title': candidate.get('Job Title') or '',
                                'Company Name': candidate.get('Company Name') or '',
                                'Job Skills': candidate.get('Job Skills') or '',
                                'Job Description': candidate.get('Job Description') or '',
                                'Job Link': candidate.get('Job Link') or '',
                                'sheet_id': candidate.get('sheet_id') or None
                            })
        return {'sheet_id': sheet_id, 'jobs': jobs}

    return {'sheet_id': None, 'jobs': []}


def style_excel_headers(file_path, retries=3, delay=0.3):
    """
    Style headers using openpyxl with retry (file lock safe-ish).
    """
    for attempt in range(retries):
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            if ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            # auto width (simple)
            for column in ws.columns:
                max_length = 0
                col_letter = None
                for cell in column:
                    if not col_letter:
                        try:
                            col_letter = cell.column_letter
                        except Exception:
                            col_letter = None
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                if col_letter:
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 60)
            wb.save(file_path)
            wb.close()
            return True
        except Exception:
            time.sleep(delay)
    return False


def ensure_primary_sheet():
    """Ensure the primary applications sheet exists and return its path."""
    path = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Address', 'Email', 'Role', 'Experience',
                   'Location', 'Github', 'LinkedIn', 'Personal Sheet Path', 'Created At']
        ws.append(headers)
        # style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        wb.save(path)
        wb.close()
    return path


# -------------------------
# Routes
# -------------------------
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy', 'service': 'Job Application Excel API'}), 200


@app.route('/api/create-personal-sheet', methods=['POST'])
def create_personal_sheet():
    """
    Create a new personal job tracking sheet using openpyxl directly.
    Body: { "name": "Alice" }
    Returns sheet_id and filename.
    """
    try:
        data = request.get_json(silent=True) or {}
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                data = {}
        name = data.get('name') or data.get('Name') or 'Unknown'
        safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '-', '_')).strip().replace(' ', '_')
        sheet_id = f"{safe_name}_{uuid.uuid4().hex[:8]}"
        filename = f"{sheet_id}_jobs.xlsx"
        filepath = os.path.join(EXCEL_DIR, filename)

        wb = Workbook()
        ws = wb.active
        headers = ['Job Title', 'Company Name', 'Job Skills',
                   'Job Description', 'Job Link', 'Updated Resume Link',
                   'Applied Date', 'Status']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # reasonable column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20

        wb.save(filepath)
        wb.close()

        return jsonify({'success': True, 'sheet_id': sheet_id, 'filename': filename, 'filepath': filepath}), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



def parse_request_data():
    """
    Parse JSON payload from Flask request.
    Supports:
      - n8n array of items (each having a .json object)
      - direct {"sheet_id": "...", "jobs": [ {...}, ... ]}
      - or single flat job with sheet_id
    """
    data = request.get_json(silent=True)
    if not data:
        return {}

    # n8n sends an array of items
    if isinstance(data, list):
        jobs = []
        sheet_id = None
        for item in data:
            # each n8n item looks like {"json": {...}}
            payload = item.get("json") if isinstance(item, dict) else {}
            if not payload:
                continue
            jobs.append(payload)
            if not sheet_id:
                sheet_id = payload.get("sheet_id")
        return {"sheet_id": sheet_id, "jobs": jobs}

    # Normal JSON body with sheet_id and jobs
    if isinstance(data, dict):
        if "jobs" in data:
            return data
        elif "sheet_id" in data:
            return {"sheet_id": data["sheet_id"], "jobs": [data]}
    
    return {}



@app.route('/api/append-jobs', methods=['POST'])
def append_jobs():
    """
    Append job listings to a personal sheet
    Accepts:
      - n8n array of items (each with job fields and sheet_id)
      - {"sheet_id":"...", "jobs":[ {...}, ... ]}
    """
    try:
        parsed = parse_request_data()
        sheet_id = parsed.get('sheet_id')
        jobs = parsed.get('jobs', [])

        # If sheet_id not provided at top-level, see if first job has it
        if not sheet_id and jobs:
            sheet_id = jobs[0].get('sheet_id')

        if not sheet_id:
            return jsonify({'success': False, 'error': 'sheet_id is required (top-level or per-item)'}), 400

        filename = f"{sheet_id.strip()}_jobs.xlsx"

        filepath = os.path.join(EXCEL_DIR, filename)
        print(EXCEL_DIR)
        print(sheet_id)
        print(filepath)
        print(os.path.exists(filepath))
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': f'Sheet not found: {filename}. Create it via /api/create-personal-sheet first.'}), 404

        # Try to load workbook with retries in case of transient lock
        retries = 3
        for attempt in range(retries):
            try:
                wb = load_workbook(filepath)
                ws = wb.active
                break
            except Exception as e:
                if attempt < retries - 1:
                    time.sleep(0.2 * (attempt + 1))
                    continue
                else:
                    return jsonify({'success': False, 'error': f'Failed to open workbook: {str(e)}'}), 500

        jobs_added = 0
        for job in jobs:
            # normalize each job
            title = job.get('Job Title') or job.get('job_title') or ''
            company = job.get('Company Name') or job.get('company') or ''
            skills = job.get('Job Skills') or job.get('job_skills') or ''
            desc = job.get('Job Description') or job.get('description') or ''
            link = job.get('Job Link') or job.get('job_link') or ''
            row = [
                title,
                company,
                skills,
                desc,
                link,
                '',  # Updated Resume Link
                datetime.now().strftime('%Y-%m-%d'),
                'Pending'
            ]
            ws.append(row)
            jobs_added += 1

        wb.save(filepath)
        wb.close()

        # style headers (best-effort)
        style_excel_headers(filepath)

        return jsonify({'success': True, 'message': 'Jobs appended', 'jobs_added': jobs_added, 'sheet_id': sheet_id, 'filepath': filepath}), 200

    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': str(e), 'traceback': traceback.format_exc()}), 500


@app.route('/api/get-jobs', methods=['GET','POST'])
def get_jobs():
    """
    Get all jobs from the shared job_data.xlsx file
    No body parameters needed
    """
    try:
        filename = "job_data.xlsx"
        filepath = os.path.join(EXCEL_DIR, filename)
        
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'Job data file not found'}), 404
        
        df = pd.read_excel(filepath)
        jobs = df.to_dict('records')
        
        return jsonify({
            'success': True, 
            'count': len(jobs), 
            'jobs': jobs
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route('/api/update-job-resume', methods=['POST'])
def update_job_resume():
    """
    Update specific job row with a resume link (identified by job_link).
    Body: { "sheet_id": "...", "job_link": "...", "resume_link": "..." }
    """
    try:
        data = request.get_json(silent=True) or {}
        sheet_id = data.get('sheet_id')
        job_link = data.get('job_link') or data.get('Job Link') or data.get('jobLink')
        resume_link = data.get('resume_link') or data.get('resumeLink') or data.get('Updated Resume Link')
        if not sheet_id or not job_link or not resume_link:
            return jsonify({'success': False, 'error': 'sheet_id, job_link and resume_link are required'}), 400
        filename = f"{sheet_id}_jobs.xlsx"
        filepath = os.path.join(EXCEL_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'Sheet not found'}), 404
        df = pd.read_excel(filepath)
        mask = df['Job Link'].astype(str) == str(job_link)
        if not mask.any():
            return jsonify({'success': False, 'error': 'Job not found'}), 404
        df.loc[mask, 'Updated Resume Link'] = resume_link
        df.loc[mask, 'Status'] = 'Resume Tailored'
        df.to_excel(filepath, index=False)
        style_excel_headers(filepath)
        return jsonify({'success': True, 'updated_rows': int(mask.sum()), 'filepath': filepath}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/primary-record', methods=['POST'])
def create_primary_record():
    """
    Add or update a primary record in primary_applications.xlsx
    Body: { "Name": "...", "Email": "...", ... }
    """
    try:
        data = request.get_json(silent=True) or {}
        primary_path = ensure_primary_sheet()
        df = pd.read_excel(primary_path)

        name = data.get('Name') or ''
        email = data.get('Email') or ''

        # find existing index by Name or Email
        existing_idx = None
        if not df.empty:
            if name and 'Name' in df.columns:
                hits = df[df['Name'].astype(str) == str(name)]
                if not hits.empty:
                    existing_idx = hits.index[0]
            if existing_idx is None and email and 'Email' in df.columns:
                hits = df[df['Email'].astype(str) == str(email)]
                if not hits.empty:
                    existing_idx = hits.index[0]

        record = {
            'Name': name,
            'Address': data.get('Address', ''),
            'Email': email,
            'Role': data.get('Role', ''),
            'Experience': data.get('Experience', ''),
            'Location': data.get('Location', ''),
            'Github': data.get('Github', ''),
            'LinkedIn': data.get('LinkedIn', ''),
            'Personal Sheet Path': data.get('Personal Sheet Path', ''),
            'Created At': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        if existing_idx is not None:
            for k, v in record.items():
                if k != 'Created At':
                    df.at[existing_idx, k] = v
            action = 'updated'
        else:
            df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
            action = 'created'

        df.to_excel(primary_path, index=False)
        style_excel_headers(primary_path)
        return jsonify({'success': True, 'action': action, 'record': record, 'file_path': primary_path}), 200

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/list-sheets', methods=['GET'])
def list_sheets():
    try:
        files = [f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]
        return jsonify({'success': True, 'files': files, 'count': len(files)}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download/<sheet_id>', methods=['GET'])
def download_sheet(sheet_id):
    filename = f"{sheet_id}_jobs.xlsx"
    filepath = os.path.join(EXCEL_DIR, filename)
    if not os.path.exists(filepath):
        if sheet_id == 'primary':
            filepath = os.path.join(EXCEL_DIR, PRIMARY_SHEET)
            filename = PRIMARY_SHEET
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'error': 'File not found'}), 404
        else:
            return jsonify({'success': False, 'error': 'File not found'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    # For local dev only. Run with gunicorn/uvicorn for production.
    app.run(host='0.0.0.0', port=5002, debug=True)
