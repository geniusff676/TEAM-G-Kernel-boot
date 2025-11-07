from flask import Flask, request, jsonify
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
import threading

app = Flask(__name__)

# Lock for thread-safe file operations
file_lock = threading.Lock()

# Configuration
EXCEL_FILE = 'job_data.xlsx'
SHEET_NAME = 'Jobs'

def initialize_excel_file():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        
        # Add headers
        headers = ['Job Title', 'Company Name', 'Job Skills', 
                   'Job Description', 'Job Link', 'Timestamp']
        ws.append(headers)
        
        # Set column widths for better readability
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 20
        
        wb.save(EXCEL_FILE)

def append_to_excel(data):
    """Safely append data to Excel file"""
    with file_lock:
        try:
            # Load existing workbook
            wb = openpyxl.load_workbook(EXCEL_FILE)
            
            # Get or create sheet
            if SHEET_NAME in wb.sheetnames:
                ws = wb[SHEET_NAME]
            else:
                ws = wb.create_sheet(SHEET_NAME)
                # Add headers if new sheet
                headers = ['Job Title', 'Company Name', 'Job Skills', 
                          'Job Description', 'Job Link', 'Timestamp']
                ws.append(headers)
            
            # Add timestamp
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Append row
            row_data = [
                data.get('Job Title', ''),
                data.get('Company Name', ''),
                data.get('Job Skills', ''),
                data.get('Job Description', ''),
                data.get('Job Link', ''),
                timestamp
            ]
            ws.append(row_data)
            
            # Save workbook
            wb.save(EXCEL_FILE)
            wb.close()
            
            return True, "Data saved successfully"
        except Exception as e:
            return False, f"Error saving data: {str(e)}"

@app.route('/store-job-data', methods=['POST'])
def store_job_data():
    """
    Endpoint to receive job data from n8n and store in Excel
    Expects JSON body with structure:
    {
        "sheet_id": "string",
        "data": [
            {
                "Job Title": "string",
                "Company Name": "string",
                "Job Skills": "string",
                "Job Description": "string",
                "Job Link": "string"
            }
        ]
    }
    """
    try:
        # Get JSON data from request
        payload = request.get_json()
        
        if not payload:
            return jsonify({
                'success': False,
                'error': 'No JSON data received'
            }), 400
        
        # Extract data array (handle both direct array and nested structure)
        if isinstance(payload, list):
            data_array = payload
        elif 'data' in payload:
            data_array = payload['data']
        else:
            # Try to find array in payload
            data_array = None
            for key, value in payload.items():
                if isinstance(value, list):
                    data_array = value
                    break
            
            if not data_array:
                return jsonify({
                    'success': False,
                    'error': 'No data array found in payload'
                }), 400
        
        # Validate data array
        if not isinstance(data_array, list):
            return jsonify({
                'success': False,
                'error': 'Data must be an array'
            }), 400
        
        # Initialize Excel file if needed
        initialize_excel_file()
        
        # Process each job entry
        results = []
        for idx, job_data in enumerate(data_array):
            if not isinstance(job_data, dict):
                results.append({
                    'index': idx,
                    'success': False,
                    'error': 'Invalid data format'
                })
                continue
            
            success, message = append_to_excel(job_data)
            results.append({
                'index': idx,
                'success': success,
                'message': message,
                'job_title': job_data.get('Job Title', 'N/A')
            })
        
        # Check if all succeeded
        all_success = all(r['success'] for r in results)
        
        return jsonify({
            'success': all_success,
            'total_records': len(data_array),
            'results': results,
            'file': EXCEL_FILE
        }), 200 if all_success else 207
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    file_exists = os.path.exists(EXCEL_FILE)
    return jsonify({
        'status': 'healthy',
        'excel_file': EXCEL_FILE,
        'file_exists': file_exists
    }), 200
import os
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
import pandas as pd

# Configuration
UPLOAD_FOLDER = 'uploads/resumes'
EXCEL_DIR = 'data'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'doc', 'docx'}

# Create directories if they don't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload', methods=['POST'])
def upload():
    """
    Upload resume file and save job data
    Expects:
    - file: resume file
    - jobtitle: job title
    - companyName: company name
    - sheet_id: sheet identifier
    - joburl: job posting URL
    """
    try:
        # Get form data
        jobtitle = request.form.get('jobtitle')
        companyName = request.form.get('companyName')
        sheet_id = request.form.get('sheet_id', 'job_data')
        joburl = request.form.get('joburl')
        
        # Validate required fields
        if not all([jobtitle, companyName, joburl]):
            return jsonify({
                'success': False, 
                'error': 'Missing required fields: jobtitle, companyName, joburl'
            }), 400
        
        # Handle file upload
        resume_filename = None
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename and allowed_file(file.filename):
                # Create safe filename
                safe_filename = secure_filename(f"{companyName}_{jobtitle}_{file.filename}")
                filepath = os.path.join(UPLOAD_FOLDER, safe_filename)
                file.save(filepath)
                resume_filename = safe_filename
        
        # Save to Excel
        excel_file = os.path.join(EXCEL_DIR, f"{sheet_id}.xlsx")
        
        # Create new job record
        job_data = {
            'job_title': jobtitle,
            'company_name': companyName,
            'job_url': joburl,
            'resume_file': resume_filename,
            'sheet_id': sheet_id
        }
        
        # Append to existing or create new Excel file
        if os.path.exists(excel_file):
            df = pd.read_excel(excel_file)
            df = pd.concat([df, pd.DataFrame([job_data])], ignore_index=True)
        else:
            df = pd.DataFrame([job_data])
        
        df.to_excel(excel_file, index=False)
        
        return jsonify({
            'success': True,
            'message': 'Job data and resume uploaded successfully',
            'job_data': job_data,
            'resume_saved': resume_filename is not None
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
    
if __name__ == '__main__':
    # Initialize Excel file on startup
    initialize_excel_file()
    
    # Run Flask app
    app.run(host='0.0.0.0', port=5000, debug=True)